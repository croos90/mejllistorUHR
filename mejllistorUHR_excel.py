import openpyxl as pxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import re
import time
import sys

# This regex pattern matches email adresses that adhere to standard conventions
regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'

# Patterns that suggest an OCR error
common_err = [r'[a-zA-Z._-]1[a-zA-Z@._-]', r'[0-9]l[0-9@]', 
              r'[a-zA-Z]11', "O", "S", "maii", "1lr", "lr1"]

# Globals to keep track
autocorrected = 0
deleted = 0
regex_errors = 0
correct = 0
errors = []
line = 0

def main(file):
    global line, autocorrected, regex_errors, correct, deleted
    start = time.time()

    # Read original excel document
    wb_original = pxl.load_workbook(file)
    ws_original = wb_original.active
    max_row = ws_original.max_row

    # Create new workbook for output
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Results"
    ws0.column_dimensions['A'].width = 25
    ws1 = wb.create_sheet("Addresses")
    ws1.column_dimensions['A'].width = 45
    ws2 = wb.create_sheet("Possible OCR-errors")
    ws2.column_dimensions['A'].width = 45

    ws1_row, ws2_row = 1,1

    for i in range(1,max_row):
        cell = ws_original.cell(row=i,column=1)
        val = cell.value
        
        # Catch None values and possible errors
        if not isinstance(val,str):
            if isinstance(val, None):
                deleted += 1
                continue
            errors.append((i,val))
            deleted += 1
            continue

        # Get rid of "Sida ..."
        if re.search("Sida ", val):
            deleted += 1
            continue

        # Get rid of whitespace
        val_stripped = val.replace(' ','')

        # Fix common OCR errors
        val_repl, c = replace_obvious_OCR_errors(val_stripped)
        autocorrected += c
        
        # Check if adress valid
        (code,email) = check(val_repl)

        # Write adress to correct sheet, regex-errors colored
        if code == 0:
            ws1.cell(row=ws1_row, column=1,value=email)
            correct += 1
            ws1_row += 1
        elif code == 1:
            ws2.cell(row=ws2_row,column=1,value=email)
            ws2_row += 1
        elif code == 2:
            cell = ws1.cell(row=ws1_row,column=1)
            cell.value = email
            cell.fill = PatternFill("solid", fgColor="00FFCC00")
            regex_errors += 1
            ws1_row += 1
    
    # Write results on first sheet
    write_results(ws0,ws1,ws2,max_row,autocorrected,deleted,regex_errors)

    # Save workbook
    wb.save('mejllista_autofix.xlsx')

    end = time.time()
    duration = end - start

    print("\nCompleted in", round(duration,2), "seconds.\n")
    print("Lines processed: ", ws_original.max_row)
    print("Auto-corrected: ", autocorrected)
    print("Lines deleted: ", deleted)
    print("Possible OCR errors: ", ws2.max_row)
    print("Regex errors: ", regex_errors)
    print("Errors: ", errors)


def replace_obvious_OCR_errors(val):
    '''Searches for patterns in the string that are obvious OCR errors and 
    replaces them with the correct expression.'''
    corr = 0
    
    # Change '.eom' to '.com'
    if re.search(".eom", val):
        val = val.replace(".eom",".com")
        corr += 1

    # Change '.corn' to '.com'
    if re.search(".corn", val):
        val = val.replace(".corn",".com")
        corr += 1

    # Change 'gmaii' to 'gmail'
    if re.search("gmaii",val):
        val = val.replace("gmaii", "gmail")
        corr += 1

    # Change 'hotmaii' to 'hotmail'
    if re.search("hotmaii", val):
        val = val.replace("hotmaii", "hotmail")
        corr += 1
    
    # Change '1ive.se' to 'live.se'
    if re.search("1ive.se", val):
        val = val.replace("1ive.se", "live.se")
        corr += 1
    
    # # Change 'a1ex' to 'alex'
    # if re.search("a1ex",val):
    #     val = val.replace("a1ex","alex")
    #     corr += 1

    # # Correct abdu1la, abdul1a, abdu11a
    # if re.search("abdul1a",val):
    #     val = val.replace("abdul1a","abdulla")
    #     corr+=1
    # if re.search("abdu1la",val):
    #     val = val.replace("abdu1la","abdulla")
    #     corr+=1
    # if re.search("abdu11a",val):
    #     val = val.replace("abdu11a","abdulla")
    #     corr+=1
    
    # # Change 'a1bin' to 'albin'
    # if re.search("a1bin",val):
    #     val=val.replace("a1bin","albin")
    #     corr+=1
    
    return val, corr

 
def check(email):
    '''
    Checks if email adress matches regex and if adress contains any common 
    OCR error pattern.

    Returns: tuple with code and email
        0 = no regex error and no OCR error
        1 = OCR error
        2 = regex error
    '''
    if(re.fullmatch(regex, email)):
        for err in common_err:
            if re.search(err, email):
              return (1, email)
        return(0,email)
    else:
        return(2,email)


def write_results(ws0,ws1,ws2,max_row,autocorrected,deleted,regex_errors):
    '''Writes results to excel'''
    ws0.cell(row=1,column=1,value="Results")
    ws0.cell(row=2,column=1,value="Processed lines")
    ws0.cell(row=2,column=2,value=max_row)
    ws0.cell(row=3,column=1,value="Autocorrected:")
    ws0.cell(row=3,column=2,value=autocorrected)
    ws0.cell(row=4,column=1,value="Deleted:")
    ws0.cell(row=4,column=2,value=deleted)
    ws0.cell(row=5,column=1,value="Possible OCR-errors:")
    ws0.cell(row=5,column=2,value=ws2.max_row)
    ws0.cell(row=6,column=1,value="Regex-errors detected:")
    ws0.cell(row=6,column=2,value=regex_errors)


if __name__=="__main__":
    main(sys.argv[1])