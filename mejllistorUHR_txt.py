from openpyxl.styles import PatternFill
from openpyxl import Workbook
import re
import time
import sys

# This regex pattern matches email adresses that adhere to standard conventions
regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'

# Patterns that suggest an OCR error
common_err = [r'[a-zA-Z._-]1[a-zA-Z@._-]', r'[0-9]l[0-9@]', 
              r'[a-zA-Z]0[a-zA-Z]', r'[a-zA-Z]11', r'0[a-zA-Z.]', 
              r'0[a-zA-Z.]', "O", "S", "maii", "1lr", "lr1"]

# Globals to keep track
autocorrected = 0
deleted = 0
regex_errors = 0
OCR_errors = 0
correct = 0
errors = []
line = 0

def main(file):
    global line, autocorrected, regex_errors, correct, deleted, OCR_errors
    start = time.time()

    # Create new workbook for output
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Results"
    ws0.column_dimensions['A'].width = 25
    ws1 = wb.create_sheet("Addresses")
    ws1.column_dimensions['A'].width = 45

    row = 1

    with open(file, 'rb') as infile:
        for l in infile.readlines():
            val = l.decode('utf-8').strip()
            line += 1

            # Catch None values and possible errors
            if not isinstance(val,str):
                if isinstance(val, None):
                    deleted += 1
                    continue
                errors.append((line,val))
                deleted += 1
                continue

            # Get rid of "Sida ..." and "EPOSTADRESS"
            if re.search("Sida ", val):
                deleted += 1
                continue
            if re.search("EPOSTADRESS", val):
                deleted += 1
                continue

            # Get rid of whitespace
            val_stripped = val.replace(' ','')

            # Fix common OCR errors
            val_repl, c = replace_obvious_OCR_errors(val_stripped)
            autocorrected += c
            
            # Check if adress valid
            (code,email) = check(val_repl)

            # Write adress to correct sheet, regex-errors colored
            if code == 0:
                ws1.cell(row=row, column=1,value=email)
                correct += 1
            elif code == 1:
                cell = ws1.cell(row=row,column=1)
                cell.value = email
                cell.fill = PatternFill("solid", fgColor="00FFFF99")
                OCR_errors += 1
                ws1.cell(row=row, column=2,value="OCR")
            elif code == 2:
                if val == '':
                    deleted += 1
                    continue
                cell = ws1.cell(row=row,column=1)
                cell.value = email
                cell.fill = PatternFill("solid", fgColor="00FFCC00")
                regex_errors += 1
                ws1.cell(row=row, column=2,value="regex")

            ws1.cell(row=row, column=3,value=row)
            row += 1
    
    # Write results on first sheet
    write_results(ws0,line,autocorrected,deleted,regex_errors)

    # Save workbook
    wb.save(file[:-4] + '_cleaned.xlsx')

    end = time.time()
    duration = end - start

    print("\nCompleted in", round(duration,2), "seconds.\n")
    print("Lines processed: ", line)
    print("Auto-corrected: ", autocorrected)
    print("Lines deleted: ", deleted)
    print("Possible OCR errors: ", OCR_errors)
    print("Regex errors: ", regex_errors)
    print("Errors: ", errors)



def replace_obvious_OCR_errors(val):
    '''
    Searches for patterns in the string that are obvious OCR errors and 
    replaces them with the correct expression.
    '''
    corr = 0
    
    # Change '.eom' to '.com'
    if re.search(".eom", val):
        val = val.replace(".eom",".com")
        corr += 1

    # Change '.corn' to '.com'
    if re.search(".corn", val):
        val = val.replace(".corn",".com")
        corr += 1

    # Change 'gmaii' to 'gmail'
    if re.search("gmaii",val):
        val = val.replace("gmaii", "gmail")
        corr += 1
    
    # Change 'grnail' to 'gmail'
    if re.search("grnail",val):
        val = val.replace("grnail","gmail")
        corr += 1

    # Change 'hotmaii' to 'hotmail'
    if re.search("hotmaii", val):
        val = val.replace("hotmaii", "hotmail")
        corr += 1
    
    # Change '1ive.se' to 'live.se'
    if re.search("1ive.se", val):
        val = val.replace("1ive.se", "live.se")
        corr += 1

    # Change 'maiLcom' to 'mail.com'
    if re.search("maiLcom",val):
        val = val.replace("maiLcom","mail.com")
        corr += 1

    # Change 'mail.cam' to 'mail.com'
    if re.search("mail.cam",val):
        val = val.replace("mail.cam","mail.com")
        corr += 1
    
    # Fix misspellings of outlook due to 'o' being interpreted as 'a'
    if re.search("autlaak",val):
        val = val.replace("autlaak","outlook")
        corr += 1
    if re.search("autlaok",val):
        val = val.replace("autlaok","outlook")
        corr += 1
    if re.search("autlook",val):
        val = val.replace("autlook","outlook")
        corr += 1
    if re.search("autloak",val):
        val = val.replace("autloak","outlook")
        corr += 1
    if re.search("outlaok",val):
        val = val.replace("outlaok","outlook")
        corr += 1
    if re.search("outloak",val):
        val = val.replace("outloak","outlook")
        corr += 1
    if re.search("outlaak",val):
        val = val.replace("outlaok","outlook")
        corr += 1

    # Change 'hatmail' to 'hotmail'
    if re.search("hatmail",val):
        val = val.replace("hatmail","hotmail")
        corr += 1
    
    return val, corr

 
def check(email):
    '''
    Checks if email adress matches regex and if adress contains any common 
    OCR error pattern.

    Returns: tuple with code and email
        0 = no regex error or OCR error
        1 = OCR error
        2 = regex error
    '''
    if(re.fullmatch(regex, email)):
        for err in common_err:
            if re.search(err, email):
              return (1, email)
        return(0,email)
    else:
        return(2,email)


def write_results(ws0,max_row,autocorrected,deleted,regex_errors):
    '''Writes results to excel'''
    ws0.cell(row=1,column=1,value="Results")
    ws0.cell(row=2,column=1,value="Processed lines")
    ws0.cell(row=2,column=2,value=max_row)
    ws0.cell(row=3,column=1,value="Autocorrected:")
    ws0.cell(row=3,column=2,value=autocorrected)
    ws0.cell(row=4,column=1,value="Deleted:")
    ws0.cell(row=4,column=2,value=deleted)
    ws0.cell(row=5,column=1,value="Possible OCR-errors:")
    ws0.cell(row=5,column=2,value=OCR_errors)
    ws0.cell(row=6,column=1,value="Regex-errors detected:")
    ws0.cell(row=6,column=2,value=regex_errors)


if __name__=="__main__":
    main(sys.argv[1])